import openpyxl
import os

files = [
    "LAP_Personal.xlsx",
    "LAP_Business.xlsx"
]

for file in files:
    if os.path.exists(file):
        out_filename = f"scratch/{file.replace(' ', '_')}_fields.txt"
        print(f"Dumping fields of {file} to {out_filename}...")
        wb = openpyxl.load_workbook(file, data_only=True)
        with open(out_filename, "w", encoding="utf-8") as f:
            f.write(f"File: {file}\n")
            f.write(f"Sheets: {wb.sheetnames}\n\n")
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                f.write(f"=================== Sheet: {sheet_name} ===================\n")
                for row in sheet.iter_rows(values_only=True):
                    if any(row):
                        vals = [str(x) if x is not None else "" for x in row]
                        f.write("\t|\t".join(vals) + "\n")
                f.write("\n\n")
    else:
        print(f"File not found: {file}")
